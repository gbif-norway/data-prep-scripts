# %%
from datetime import datetime
import logging
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from uuid import uuid4

logging.basicConfig(level=logging.INFO)


class Biorehab:
    def __init__(self, input_file) -> None:
        self.input_file = Path(input_file)
        file_name = self.input_file.stem
        self.output_dir = Path('output')
        self.orig_with_uuid_file = self.output_dir / f'{file_name}_uuid.xlsx'
        self.ready_for_dwc_file = self.output_dir / f'{file_name}_dwc.xlsx'

    def add_missing_occurrence_ids(self):
        # Add uuid to the occurrenceId column if missing. Save as orig_with_uuid_file
        wb = load_workbook(self.input_file)
        ws = wb.active
        row_id = 3
        record_id = ws[f'B{row_id}']
        while record_id.value != None:
            occurrence_id_cell = ws[f'A{row_id}']
            if occurrence_id_cell.value == None:
                ws[f'A{row_id}'] = str(uuid4())
            row_id += 1
            record_id = ws[f'B{row_id}']
        wb.save(self.orig_with_uuid_file)

    def trim_for_dwc_publication(self):
        try:
            data = pd.read_excel(self.orig_with_uuid_file, skiprows=[1])
        except Exception as e:
            raise Exception(
                f'Unable to read {self.orig_with_uuid_file}. Did you run \
                    `add_missing_occurrence_ids()` first? \nTraceback: {e}')
        dwc_columns = [x for x in data.columns if 'Unnamed:' not in x]
        data = data[dwc_columns]
        data.to_excel(self.ready_for_dwc_file)
        # data['dateIdentified'] = data.apply(_get_date_identified, axis=1)

    # made because a missed the column - might still be needed
    # def _get_date_identified(row: pd.Series):
    #     try:
    #         if not type(row['yearIdentified']) is str:
    #             # logging.debug('No identified date')
    #             return None
    #         date = datetime(year=int(row['yearIdentified']),
    #                         month=int(row['monthIdentified']),
    #                         day=int(row['dayIdentified']))
    #         return date.strftime('%Y-%m-%d')
    #     except:
    #         raise Exception(f'Unable to get identifiedDate: {row}')
