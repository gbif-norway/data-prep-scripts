#%%
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from uuid import uuid4

# %%
input_file = Path('input') / 'Biorehab Klepsland_Norge_1.xlsx'
output_dir = Path('output')
orig_with_uuid_file = output_dir / 'Biorehab Klepsland_Norge_1_uuid.xlsx'
ready_for_dwc_file = output_dir / 'Biorehab Klepsland_Norge_1_dwc.xlsx'
# %% Add uuid to the occurrenceId column if missing. Save as orig_with_uuid_file
wb = load_workbook(input_file) 
ws = wb.active 
row_id = 3
record_id = ws[f'B{row_id}']
while record_id.value != None:
    occurrence_id_cell = ws[f'A{row_id}']
    if occurrence_id_cell.value == None:
        ws[f'A{row_id}'] = str(uuid4())
    row_id += 1
    record_id = ws[f'B{row_id}']
wb.save(orig_with_uuid_file) 
# %% drop unused columns for dwc
data = pd.read_excel(orig_with_uuid_file, skiprows=[1]) 

# %%
dwc_columns = [x for x in data.columns if 'Unnamed:' not in x]
data = data[dwc_columns]
data.to_excel(ready_for_dwc_file)
# %%
# %%
